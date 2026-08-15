# -*- coding: utf-8 -*-
"""Escreve o calendário operacional em XLSX (3 abas):
  Premissas  — parâmetros usados (células de entrada destacadas; para
               recalcular a alocação é preciso rodar o pipeline de novo)
  Calendário — uma linha por turno por ponto (16/08 a 03/10)
  Resumo     — totais por fase e por ponto via SUMIFS/COUNTIFS (recalculam
               se o usuário editar a coluna de panfletos do Calendário)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import plano_config

ARIAL = Font(name="Arial", size=10)
NEGRITO = Font(name="Arial", size=10, bold=True)
TITULO = Font(name="Arial", size=13, bold=True)
AZUL_INPUT = Font(name="Arial", size=10, color="0000FF")
AMARELO = PatternFill("solid", fgColor="FFFF00")
CINZA = PatternFill("solid", fgColor="DDDDDD")


def _cabecalho(ws, linha, textos):
    for j, t in enumerate(textos, 1):
        c = ws.cell(row=linha, column=j, value=t)
        c.font = NEGRITO
        c.fill = CINZA


def escrever_xlsx(cal: pd.DataFrame, caminho):
    wb = Workbook()

    # ---------------- Premissas ----------------
    ws = wb.active
    ws.title = "Premissas"
    ws["A1"] = "Plano de panfletagem — Florianópolis 2026 (Matheus Cadorin)"
    ws["A1"].font = TITULO
    ws["A3"] = ("PREMISSAS OPERACIONAIS — valores em amarelo NÃO foram informados "
                "pelo coordenador e são premissas de trabalho. Para recalcular a "
                "alocação com outros valores, edite pipeline/plano_config.py e rode "
                "python3 run_plano.py (a planilha é gerada de novo).")
    ws["A3"].font = ARIAL
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A3:E5")
    premissas = [
        ("Equipe fixa em dia útil (pessoas)", plano_config.EQUIPE["fixos_dia_util"]),
        ("Equipe no sábado (fixos + voluntários)", plano_config.EQUIPE["sabado"]),
        ("Equipe no domingo (ação leve opcional)", plano_config.EQUIPE["domingo"]),
        ("Orçamento total de panfletos (tiragem)", plano_config.ORCAMENTO_PANFLETOS),
        ("Fator aplicado p/ caber no orçamento", cal.attrs.get("fator_orcamento", 1.0)),
        ("Período", "16/08/2026 a 03/10/2026 (dia 04/10 excluído — boca de urna é crime)"),
        ("Taxa de aceite veículo / pedestre / feira (estimativas)",
         "45% / 18% / 40%"),
        ("Tempo por entrega em semáforo (estimativa)", "6 segundos"),
    ]
    _cabecalho(ws, 7, ["Parâmetro", "Valor"])
    for i, (k, v) in enumerate(premissas, 8):
        ws.cell(row=i, column=1, value=k).font = ARIAL
        c = ws.cell(row=i, column=2, value=v)
        c.font = AZUL_INPUT
        if i <= 11:
            c.fill = AMARELO
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 60

    # ---------------- Calendário ----------------
    ws = wb.create_sheet("Calendário")
    cols = ["data", "dia", "fase", "turno", "inicio", "fim", "prioridade",
            "ponto", "endereco", "regiao", "tipo", "pessoas", "panfletos",
            "justificativa", "alternativa_chuva", "observacao"]
    rotulos = ["Data", "Dia", "Fase", "Turno", "Início", "Fim", "Prior.",
               "Ponto", "Endereço", "Região", "Tipo", "Pessoas", "Panfletos",
               "Justificativa", "Alternativa se chover", "Observação"]
    _cabecalho(ws, 1, rotulos)
    for i, (_, r) in enumerate(cal.iterrows(), 2):
        for j, c in enumerate(cols, 1):
            cel = ws.cell(row=i, column=j, value=r[c])
            cel.font = ARIAL
    # colunas: A=data B=dia C=fase D=turno E=início F=fim G=prior. H=ponto
    #          I=endereço J=região K=tipo L=pessoas M=panfletos ...
    n = len(cal) + 1
    tot = ws.cell(row=n + 1, column=12, value="TOTAL")
    tot.font = NEGRITO
    f = ws.cell(row=n + 1, column=13, value=f"=SUM(M2:M{n})")
    f.font = NEGRITO
    ws.freeze_panes = "A2"
    larguras = [11, 5, 34, 9, 7, 7, 7, 52, 40, 14, 12, 8, 10, 62, 42, 55]
    for j, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---------------- Resumo ----------------
    ws = wb.create_sheet("Resumo")
    ws["A1"] = "Resumo por fase"
    ws["A1"].font = TITULO
    _cabecalho(ws, 3, ["Fase", "Turnos", "Panfletos"])
    fases = list(dict.fromkeys(cal["fase"]))
    for i, fase in enumerate(fases, 4):
        ws.cell(row=i, column=1, value=fase).font = ARIAL
        ws.cell(row=i, column=2,
                value=f'=COUNTIF(Calendário!C:C,A{i})').font = ARIAL
        ws.cell(row=i, column=3,
                value=f'=SUMIF(Calendário!C:C,A{i},Calendário!M:M)').font = ARIAL

    lin = len(fases) + 6
    ws.cell(row=lin, column=1, value="Pontos mais acionados").font = TITULO
    _cabecalho(ws, lin + 2, ["Ponto", "Turnos", "Panfletos"])
    tops = cal.groupby("ponto")["panfletos"].sum().nlargest(15)
    for i, ponto in enumerate(tops.index, lin + 3):
        ws.cell(row=i, column=1, value=ponto).font = ARIAL
        ws.cell(row=i, column=2,
                value=f'=COUNTIF(Calendário!H:H,A{i})').font = ARIAL
        ws.cell(row=i, column=3,
                value=f'=SUMIF(Calendário!H:H,A{i},Calendário!M:M)').font = ARIAL
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12

    wb.save(caminho)
    print(f"  calendário salvo em {caminho}")
